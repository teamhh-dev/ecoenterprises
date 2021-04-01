from Summary import *
from openpyxl import *
import datetime
# from HomePageView import *
from copy import *
from Bill import *
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import *


class ExcelModel():
    def __init__(self):

        # all global variables
        self.workbook = None
        self.worksheet = None
        self.activeWorkSheet = None
        self.fileName = None
        self.firstRow = 0
        self.service_len = 0
        self.sr_number = 0
        self.chargesQtyCell = None
        self.visitTypeCell = None
        self.visitRatecell = None
    #  Selecting Excel file in which data has to stored

    def selectExcelFile(self, zone: str):
        year = datetime.datetime.now().strftime("%Y")
        month = datetime.datetime.now().strftime("%m")
        excelFileName = "Monthly Report "+zone.upper()+" " + \
            "01-"+"{:0>2d}".format(int(month))+"-"+year+".xlsx"
        # self.fileName = 'User Data/Feburary/'+zone+'.xlsx'
        monthFullName = datetime.datetime.now().strftime("%B")
        self.fileName = 'User Data/'+monthFullName.upper()+"_"+year + "/"+excelFileName
        print(self.fileName)
        self.workbook = load_workbook(filename=self.fileName)
        self.worksheet = self.workbook['template']

    # Creating worksheet
    def createWorksheet(self, name: str):

        self.activeWorkSheet = self.workbook.copy_worksheet(self.worksheet)
        self.activeWorkSheet.title = name

    # Adding Bill to Excel sheet
    def addBill(self, bill: Bill):

        self.selectExcelFile(bill.getComplainInfo().getZone())
        self.createWorksheet(bill.getComplainInfo().getInvoiceId())
        self.saveBillComlaintInfo(bill.getComplainInfo())
        self.saveBillServices(bill.getServices())
        self.saveAdditionalCharges(bill.getAdditionalCharges())

        self.workbook.save(filename=self.fileName)

    # Adding essential information about complaint in the worksheet
    def saveBillComlaintInfo(self, complainInfo: ComplaintInfo):

        self.activeWorkSheet['D2'] = complainInfo.getInvoiceId()
        self.activeWorkSheet['A7'] = complainInfo.getDate()
        self.activeWorkSheet['H7'] = complainInfo.getComplainNo()
        self.activeWorkSheet['A8'] = complainInfo.getBankName()
        self.activeWorkSheet['A9'] = complainInfo.getAddress()

    # adding services rows to active worksheet , format them and then adding data in them
    def saveBillServices(self, services: Services):

        self.firstRow = 16

        self.service_len = len(services.getServicesList())-2
        if not self.service_len == 0 and not self.service_len == -1 and not self.service_len == -2:
            self.activeWorkSheet.insert_rows(17, self.service_len)
        row = list(self.activeWorkSheet.rows)[15]
        i = 0
        j = self.firstRow

    # Adding service rows to active worksheet
        if not self.service_len == 0 and not self.service_len == -1 and not self.service_len == -2:
            while j <= self.firstRow + self.service_len:

                for cell in row:

                    list(self.activeWorkSheet.rows)[
                        j][i].border = copy(cell.border)
                    list(self.activeWorkSheet.rows)[
                        j][i].font = copy(cell.font)
                    list(self.activeWorkSheet.rows)[
                        j][i].number_format = copy(cell.number_format)
                    i = i+1
                i = 0
                j = j+1

        services.getServicesList().reverse()
        self.sr_number = 1
        for service in services.getServicesList():

            self.activeWorkSheet['A'+str(self.firstRow)] = self.sr_number
            self.activeWorkSheet['B' +
                                 str(self.firstRow)] = service['description']
            self.activeWorkSheet['E'+str(self.firstRow)] = service['qty']
            self.activeWorkSheet['G'+str(self.firstRow)] = service['rate']
            self.activeWorkSheet['H'+str(self.firstRow)] = service['amount']
            self.firstRow = self.firstRow + 1
            self.sr_number = self.sr_number + 1
        if not self.service_len == 0 and not self.service_len == -1 and not self.service_len == -2:
            chargesFormula = "=E{rowno}*G{rowno}".format(rowno=j+3)
            self.activeWorkSheet["H"+str(j+1)] = "=SUM(H16:H"+str(j)+")"
            self.activeWorkSheet["H" + str(j+2)] = "=H" + str(j+1)+"*0.16"
            self.activeWorkSheet["H" +
                                 str(j+5)] = "=SUM(H" + str(j+1)+":H"+str(j+4)+")"
            # self.activeWorkSheet["H"+str(j+3)] = "=E" + \
            #     str(j+3)+"*"+"G"+str(j+3)
            self.activeWorkSheet["H"+str(j+3)] = chargesFormula
# Calling the formatting cells function
        self.formattingCells(16)

    def saveAdditionalCharges(self, additionalCharges: AdditionalCharges):
        print("ok")
        print(self.chargesQtyCell)
        # Perform formatting cells
        self.activeWorkSheet[self.chargesQtyCell] = additionalCharges.getPersonQty(
        )
        self.activeWorkSheet[self.visitTypeCell] = additionalCharges.getType(
        )+" Charges"
        self.activeWorkSheet[self.visitRatecell] = additionalCharges.getPerPersonRate(
        )
        print(self.chargesQtyCell)

    def formattingCells(self, rowNo: int):
        self.firstRow = 16
        while rowNo <= (self.firstRow+self.service_len+1) or rowNo == 16 or rowNo == 17:
            print(rowNo)
            self.activeWorkSheet.merge_cells('B'+str(rowNo)+':'+'D'+str(rowNo))
            self.activeWorkSheet.cell(row=rowNo, column=1).alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)
            self.activeWorkSheet.cell(row=rowNo, column=2).alignment = Alignment(
                horizontal='left', vertical='top', wrap_text=True)
            self.activeWorkSheet.cell(row=rowNo, column=5).alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)
            self.activeWorkSheet.cell(row=rowNo, column=7).alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)
            self.activeWorkSheet.cell(row=rowNo, column=8).alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)

            #  Increasing the size of height of row
            if (len(str(self.activeWorkSheet.cell(row=rowNo, column=2).value))) > 36:

                self.activeWorkSheet.row_dimensions[rowNo].height = 31.50
            # fixed
            rowNo = rowNo + 1
        self.chargesQtyCell = 'E'+str(rowNo+2)
        self.visitTypeCell = 'B'+str(rowNo+2)
        self.visitRatecell = 'G'+str(rowNo+2)

    def addToSummary(self, summary: Summary):
        file = 'User Data\Summary.xlsx'
        workbook = load_workbook(filename=file)
        activeSheet = workbook['Bill Summary']
        # activeSheet = workbook.active

        # inserting and formatting rows
        noOfRows = len(summary.getRecordsList()) - 3
        if noOfRows > 0:
            activeSheet.insert_rows(11, noOfRows)
            formattedrow = list(activeSheet.rows)[9]
            print(formattedrow[0])
            # unformattedrow = list(worksheet.rows)[9]
            for row in range(noOfRows):
                unformattedrow = list(activeSheet.rows)[10+row]
                print(unformattedrow[0])
                for (c1, c2) in zip(formattedrow, unformattedrow):
                    c2.style = copy(c1.style)
                    c2.border = copy(c1.border)
                    c2.number_format = copy(c1.number_format)
                    c2.font = copy(c1.font)
                    c2.alignment = copy(c1.alignment)
                    c2.fill = copy(c1.fill)

        # inserting data into summary
        records = summary.getRecordsList()
        totalRecords = len(summary.getRecordsList())
        record = 0
        fromRow = 10
        toRow = fromRow + totalRecords
        i = 0
        for i in range(fromRow, toRow):
            print(records[record][0])
            activeSheet['A'+str(i)] = record + 1
            activeSheet['B'+str(i)] = records[record][0]
            activeSheet['C'+str(i)] = "[" + str(records[record][1]) + "]"
            activeSheet['D'+str(i)] = records[record][2]
            activeSheet['E'+str(i)] = records[record][3]
            activeSheet['F'+str(i)] = records[record][4]
            activeSheet['G'+str(i)] = records[record][3] + \
                records[record][4]
            activeSheet['H'+str(i)] = records[record][5]
            activeSheet['J'+str(i)] = records[record][6]
            record = record + 1

        # updating formulas
        formulaCells = ['E', 'F', 'G', 'H', 'J']
        if noOfRows > 0:
            for formulaCell in formulaCells:
                activeSheet[formulaCell+str(toRow)] = "=SUM(" + \
                    formulaCell+str(fromRow)+":"+str(toRow-1)+")"

        workbook.save(filename=file)

    def addDateToBill(self, invoiceIdAndZone: str, date: str):
        print(invoiceIdAndZone, ":", date)
        year = datetime.datetime.now().strftime("%Y")
        month = datetime.datetime.now().strftime("%m")
        excelFileName = "Monthly Report "+invoiceIdAndZone["zone"].upper()+" " + \
            "01-"+"{:0>2d}".format(int(month))+"-"+year+".xlsx"
        monthFullName = datetime.datetime.now().strftime("%B")
        fileName = 'User Data/'+monthFullName.upper()+"_"+year + "/"+excelFileName
        print(fileName)
        workbook = load_workbook(
            filename=fileName)
        worksheet = workbook[invoiceIdAndZone["invoice_id"]]
        worksheet['A7'] = date
        workbook.save(
            filename=fileName)

    def deleteSheet(self, invoiceIdAndZone: dict, lastSheetName: str):
        year = datetime.datetime.now().strftime("%Y")
        month = datetime.datetime.now().strftime("%m")
        excelFileName = "Monthly Report "+invoiceIdAndZone["zone"].upper()+" " + \
            "01-"+"{:0>2d}".format(int(month))+"-"+year+".xlsx"
        monthFullName = datetime.datetime.now().strftime("%B")
        fileName = 'User Data/'+monthFullName.upper()+"_"+year + "/"+excelFileName
        print(fileName, ":",
              invoiceIdAndZone['invoice_id'], ":", lastSheetName)
        workbook = load_workbook(
            filename=fileName)
        worksheetToBeDeleted = workbook[invoiceIdAndZone["invoice_id"]]
        lastWorksheet = workbook[lastSheetName]

        offset = int(
            worksheetToBeDeleted.title[2:5])-int(lastWorksheet.title[2:5])
        if offset != 0:
            workbook.move_sheet(lastWorksheet, offset=offset)
            workbook.remove_sheet(worksheetToBeDeleted)
            lastWorksheet.title = invoiceIdAndZone['invoice_id']
            lastWorksheet['D2'] = invoiceIdAndZone['invoice_id']

        else:
            workbook.remove_sheet(worksheetToBeDeleted)

        workbook.save(filename=fileName)

        # print(lastWorksheet.title, worksheetToBeDeleted.title)
        # workbook.remove_sheet(worksheetToBeCopiedTo)
        # offset = int(worksheetToBeDeleted.title) - \
        #     int(workbook._sheets[lastSheetIdx].title[2:5])

        # workbook.move_sheet(lastWorksheet, offset=offset)

        # workbook = load_workbook(
        #     filename=fileName)

        # workbook.dele
        # print(workbook._sheets)
        # # names = workbook.get_sheet_names()[2:]
        # # names.sort()
        # # print(names)
        # # print(workbook._sheets[2:])
        # # workbook._sheets.sort(key=lambda ws: ws.title)
        # # print(workbook._sheets)
        # lastSheetIdx = len(workbook.sheetnames)-1
        # print(lastSheetIdx)
        # offset = int(workbook._sheets[3].title[2:5]) - \
        #     int(workbook._sheets[lastSheetIdx].title[2:5])
        # # print(int(offset))
        # # workbook._sheets.sort(key=lambda ws: ws.title)
        # print(workbook._sheets)

        # workbook.save(filename=fileName)
